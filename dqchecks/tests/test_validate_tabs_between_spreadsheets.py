"""
Testing validate_tabs_between_spreadsheets function from panacea
"""
import pytest
from openpyxl import Workbook
from dqchecks.panacea import validate_tabs_between_spreadsheets

@pytest.fixture
def workbook1():
    """Testing workbook with 2 sheets"""
    wb = Workbook()
    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet2")
    return wb


@pytest.fixture
def workbook2():
    """Testing workbook with 2 sheets"""
    wb = Workbook()
    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet2")
    return wb


@pytest.fixture
def workbook3():
    """Testing workbook with 2 sheets"""
    wb = Workbook()
    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet3")
    return wb


@pytest.fixture
def empty_workbook():
    """Return empty openpyxl workbook"""
    return Workbook()

# pylint: disable=W0621
def test_same_tabs(workbook1, workbook2):
    """Test case where both workbooks have the same sheet names."""
    are_same, message = validate_tabs_between_spreadsheets(workbook1, workbook2)
    assert are_same is True
    assert message == "Both spreadsheets have the same sheet names."

# pylint: disable=W0621
def test_different_tabs(workbook1, workbook3):
    """Test case where the workbooks have different sheet names."""
    are_same, message = validate_tabs_between_spreadsheets(workbook1, workbook3)
    assert are_same is False
    assert "Spreadsheet 1 is missing the following sheets:" in message
    assert "Spreadsheet 2 is missing the following sheets:" in message

# pylint: disable=W0621
def test_empty_workbook(workbook1, empty_workbook):
    """Test case where one workbook is empty (no sheets)."""
    are_same, message = validate_tabs_between_spreadsheets(workbook1, empty_workbook)
    assert are_same is False
    assert "Spreadsheet 2 is missing the following sheets:" in message


def test_invalid_type(workbook2):
    """Test case where the argument is not a valid openpyxl workbook."""
    with pytest.raises(ValueError):
        validate_tabs_between_spreadsheets("not_a_workbook", workbook2)

    with pytest.raises(ValueError):
        validate_tabs_between_spreadsheets(workbook2, "not_a_workbook")


def test_invalid_object(workbook2):
    """Test case where the argument is an invalid object."""
    with pytest.raises(ValueError):
        validate_tabs_between_spreadsheets(None, workbook2)

    with pytest.raises(ValueError):
        validate_tabs_between_spreadsheets(workbook2, None)

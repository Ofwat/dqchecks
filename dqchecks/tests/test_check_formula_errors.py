import pytest
from openpyxl import Workbook

from dqchecks.panacea import check_formula_errors

@pytest.fixture
def workbook_with_no_errors():
    """Fixture for a workbook with no formula errors."""
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = 10
    sheet['A2'] = 20
    sheet['A3'] = "=A1+A2"  # Valid formula
    return sheet

@pytest.fixture
def workbook_with_errors():
    """Fixture for a workbook with formula errors."""
    wb = Workbook()
    sheet = wb.active

    # Adding valid data
    sheet['A1'] = 10
    sheet['A2'] = 0  # Division by zero will happen in A3

    # Adding formulas that would cause errors in Excel
    sheet['A3'] = "=A1/A2"  # Division by zero (#DIV/0!)
    sheet['A4'] = "=SUM('InvalidRange')"  # Invalid range (#REF!)
    sheet['A5'] = "=A1 + 'InvalidRange'"  # Unrecognized range (#NAME?)

    # Manually setting the formulas to simulate errors in Excel
    # Openpyxl itself will not evaluate these, but in Excel they would be errors.
    sheet['A3'].value = '#DIV/0!'  # Manually simulate the error for testing purposes
    sheet['A4'].value = '#REF!'    # Manually simulate the error for testing purposes
    sheet['A5'].value = '#NAME?'   # Manually simulate the error for testing purposes

    return sheet


@pytest.fixture
def workbook_with_non_formula_cells():
    """Fixture for a workbook with non-formula cells."""
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = 10
    sheet['A2'] = 0
    sheet['A3'] = "Hello"
    return sheet

def test_check_formula_errors_no_errors(workbook_with_no_errors):
    """Test that the function returns 'No formula errors found' when there are no formula errors."""
    sheet = workbook_with_no_errors
    result, message = check_formula_errors(sheet)
    assert result is True
    assert message == "No formula errors found."

def test_check_formula_errors_with_errors(workbook_with_errors):
    """Test that the function correctly identifies formula errors."""
    sheet = workbook_with_errors
    result, message = check_formula_errors(sheet)
    assert result is False
    assert "Error in Sheet!A3: #DIV/0!" in message
    assert "Error in Sheet!A4: #REF!" in message
    assert "Error in Sheet!A5: #NAME?" in message

def test_check_formula_errors_with_non_formula_cells(workbook_with_non_formula_cells):
    """Test that non-formula cells don't affect the result."""
    sheet = workbook_with_non_formula_cells
    result, message = check_formula_errors(sheet)
    assert result is True
    assert message == "No formula errors found."

def test_check_formula_errors_invalid_input():
    """Test that the function raises a ValueError for invalid input types."""
    with pytest.raises(ValueError):
        check_formula_errors("invalid_input")  # Passing a string instead of a worksheet

def test_check_formula_errors_invalid_sheet_type():
    """Test that the function raises a ValueError when the input is not a Worksheet object."""
    class InvalidSheet:
        pass
    with pytest.raises(ValueError):
        check_formula_errors(InvalidSheet())  # Passing a non-worksheet object

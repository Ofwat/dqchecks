import pandas as pd
from openpyxl.styles import PatternFill

green_fill = PatternFill(start_color='72C931', end_color='72C931', fill_type='solid')
red_fill = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')

##Import Company excel file
def import_data(file_location, excel_error_log_name):
    firstsheet = excel_error_log_name['Intro']
    print('File Location Name : ', file_location)
    try:
        xl1 = pd.ExcelFile(file_location)  # Define the excel file
        worksheets = xl1.sheet_names  # Get the list of worksheets in the file
        found_dictionary, found_f_outputs = (False, False)
        dict_of_sheets = {}
        for sheet in worksheets:  # iterate through the sheets in the file
            if sheet.startswith("Dict_"):  # ...if the worksheet starts with the text "Dict_" then...
                found_dictionary = True
                dict_of_sheets[sheet] = (pd.read_excel(xl1, sheet_name=sheet, skiprows=[0, 2]), "Dict_")
            if sheet.startswith("fOut_"):  # ...if the workseet starts with the text "fOut" then...
                found_f_outputs = True
                dict_of_sheets[sheet] = (pd.read_excel(xl1, sheet_name=sheet, skiprows=[0, 2]), "fOut_")
        if not found_f_outputs:
            firstsheet['B6'].value = f"{firstsheet['B6'].value} Error message: A worksheet starting with 'fOut_' (an F_Outputs sheet) was not found in worksheets, Proteus has failed to run"
            firstsheet['B6'].fill = red_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError}"
        firstsheet['B6'].fill = red_fill
        pass
    return dict_of_sheets
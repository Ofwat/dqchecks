import datetime
import logging
from collections import namedtuple
from openpyxl.workbook.workbook import Workbook
import pandas as pd
import re

logging.basicConfig(level=logging.INFO)

# Define namedtuple for context
ProcessingContext = namedtuple(
    'ProcessingContext', ['org_cd', 'submission_period_cd', 'process_cd',
                          'template_version', 'last_modified']
)

def is_valid_regex(pattern: str) -> bool:
    try:
        re.compile(pattern)  # Try to compile the regex pattern
        return True  # If no exception, it's a valid regex
    except re.error:  # If an exception is raised, it's not a valid regex
        return False

def process_fout_sheets(
        wb: Workbook,
        context: ProcessingContext,
        observation_patterns: list[str],
        ):
    """
    Processes all sheets in the given Excel workbook that start with 'fOut_'.

    Args:
        wb (openpyxl.workbook.workbook.Workbook): The openpyxl Workbook object to process.
        context (ProcessingContext): The context object containing organization code,
                                     submission period code, process code, template version,
                                     and last modified timestamp.

    Returns:
        pd.DataFrame: A pandas DataFrame containing the processed data from all matching sheets.

    Raises:
        ValueError: If any of the input values are invalid or if no matching sheets are found.
        TypeError: If the provided workbook is not of the correct type.
        Exception: For any other unexpected errors.
    """

    # Input checks
    if not isinstance(wb, Workbook):
        raise TypeError("The 'wb' argument must be a valid openpyxl workbook object.")

    # Check if context fields are of valid type
    if not isinstance(context.org_cd, str) or not context.org_cd:
        raise ValueError("The 'org_cd' argument must be a non-empty string.")

    if not isinstance(context.submission_period_cd, str) or not context.submission_period_cd:
        raise ValueError("The 'submission_period_cd' argument must be a non-empty string.")

    if not isinstance(context.process_cd, str) or not context.process_cd:
        raise ValueError("The 'process_cd' argument must be a non-empty string.")

    if not isinstance(context.template_version, str) or not context.template_version:
        raise ValueError("The 'template_version' argument must be a non-empty string.")

    if not isinstance(context.last_modified, datetime.datetime):
        raise ValueError("The 'last_modified' argument must be a datetime object.")

    if (not isinstance(observation_patterns, list)
            or not all(isinstance(i, str) for i in observation_patterns)
            or not all(is_valid_regex(i) for i in observation_patterns)
            ):
        raise ValueError("The 'observation_patterns' argument needs to be a list of regex strings.")

    # Check if the workbook is in data_only mode
    if not wb.data_only:
        logging.warning("Reading in non data_only mode. Some data may not be accessible.")
        
    logging.info("Using observation patterns: %s", observation_patterns)

    # Filter sheets that start with 'fOut_'
    fout_sheets = [
        sheet for sheet in wb.sheetnames if sheet.startswith("fOut_")
    ]

    # If no matching sheets are found, raise an error
    if not fout_sheets:
        raise ValueError(f"No fOut_* sheets found. Available sheets: {wb.sheetnames}")

    # Read matching sheets into DataFrames, skip the first row, and use the second row as header
    df_list = []
    for sheetname in fout_sheets:
        data = wb[sheetname].iter_rows(min_row=2, values_only=True)
        # Get the header from the first row (if the first row contains the column names)
        try:
            headers = next(data)
        except StopIteration as exc:
            raise ValueError(f"Sheet '{sheetname}' is empty or has no data.") from exc

        # Convert the sheet to a pandas DataFrame
        df = pd.DataFrame(data, columns=headers)
        df["Sheet_Cd"] = sheetname  # Add the sheet name as a new column 'Sheet_Cd'
        df_list.append(df)

    # Drop rows that are completely NaN (ignoring the 'Sheet_Cd' column)
    df_list = [
        df.dropna(how='all', subset=df.columns.difference(['Sheet_Cd']))
        for df in df_list
    ]

    # If no valid rows remain after dropping NaN rows, raise an error
    if any(i.empty for i in df_list):
        raise ValueError("No valid data found after removing rows with NaN values.")

    # Initialize an empty list to collect the melted DataFrames
    melted_dfs = []

    # Loop through each DataFrame in df_list and process them
    for df in df_list:
        observation_period_columns = []
        # Identify columns which match observation patterns supplied
        for observation_pattern in observation_patterns:
            observation_period_columns += list(
                    df.filter(regex=observation_pattern).columns.tolist()
                )
        observation_period_columns = set(observation_period_columns)

        # If no observation period columns are found, raise an error
        if not observation_period_columns:
            raise ValueError("No observation period columns found in the data.")

        # Get ID columns (all columns except observation period columns)
        id_columns = set(df.columns.tolist()) - observation_period_columns

        # Pivot the DataFrame to melt the observation period columns into rows
        pivoted_df = df.melt(
            id_vars=id_columns,
            var_name="Observation_Period_Cd",
            value_name="Measure_Value"
        )

        # Add static columns from the context
        pivoted_df["Organisation_Cd"] = context.org_cd
        pivoted_df["Submission_Period_Cd"] = context.submission_period_cd
        pivoted_df["Process_Cd"] = context.process_cd
        pivoted_df["Template_Version"] = context.template_version
        pivoted_df["Submission_Date"] = context.last_modified  # Last modified date

        # Convert all columns to string type for consistency
        pivoted_df = pivoted_df.astype(str)

        # Define column renaming mapping
        column_rename_map = {
            'Organisation_Cd': 'Organisation_Cd',
            'Submission_Period_Cd': 'Submission_Period_Cd',
            'Observation_Period_Cd': 'Observation_Period_Cd',
            'Process_Cd': 'Process_Cd',
            'Template_Version': 'Template_Version',
            'Sheet_Cd': 'Sheet_Cd',
            'Reference': 'Measure_Cd',
            'Measure_Value': 'Measure_Value',
            'Item description': 'Measure_Desc',
            'Unit': 'Measure_Unit',
            'Model': 'Model_Cd',
            'Submission_Date': 'Submission_Date'
        }

        # Rename columns using the defined mapping
        pivoted_df = pivoted_df.rename(columns=column_rename_map)

        # Reorder columns according to the desired order
        ordered_columns = list(column_rename_map.values())
        pivoted_df = pivoted_df[ordered_columns]

        # Append the melted DataFrame to the list
        melted_dfs.append(pivoted_df)

    # Concatenate all the melted DataFrames into one big DataFrame
    final_df = pd.concat(melted_dfs, ignore_index=True)

    return final_df
